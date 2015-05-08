
from collections import OrderedDict
import csv
from datetime import datetime
import re

from babel.dates import format_date
import xlrd
import xmltodict

from mangrove.errors.MangroveException import MultipleSubmissionsForSameCodeException, SMSParserInvalidFormatException, \
    CSVParserInvalidHeaderFormatException, XlsParserInvalidHeaderFormatException, FormModelDoesNotExistsException
from mangrove.form_model.field import GeoCodeField, DateField, IntegerField, FieldSet, PhotoField, VideoField, AudioField, \
    TimeField, DateTimeField, MediaField
from mangrove.form_model.form_model import get_form_model_by_code
# from mangrove.transport.player.player import SMSPlayer
from mangrove.utils.types import is_empty, is_string
from mangrove.contrib.registration import REGISTRATION_FORM_CODE
from openpyxl import load_workbook
import StringIO



class SMSParserFactory(object):
    MESSAGE_PREFIX = ur'^(\w+)\s+\.(\w+)\s+(\w+)'

    def getSMSParser(self, message, dbm=None):
        clean_message = SMSParser(dbm).clean(message)
        if re.match(self.MESSAGE_PREFIX, clean_message, flags=re.UNICODE):
            return KeyBasedSMSParser(dbm)
        return OrderSMSParser(dbm)


class SMSParser(object):
    def __init__(self, dbm):
        self.dbm = dbm

    def _to_unicode(self, message):
        if type(message) is not unicode:
            message = unicode(message, encoding='utf-8')
        return message

    def clean(self, message):
        message = self._to_unicode(message)
        return message.strip()

    def validate_format(self, message_prefix_regex, message):
        if not re.match(message_prefix_regex, message, flags=re.UNICODE):
            raise SMSParserInvalidFormatException(message)

    def pop_form_code(self, tokens):
        form_code = tokens[0].strip().lower()
        tokens.remove(tokens[0])
        return form_code

    def get_form_code_and_tokens(self, token):
        form_code = token[0].lower()
        try:
            form_model = get_form_model_by_code(self.dbm, form_code)
            token.remove(token[0])
        except FormModelDoesNotExistsException:
            form_code = "poll"
            token = " ".join(token)
        return form_code, [token]

    def parse(self, message):
        pass

    def form_code(self, message):
        pass

    def select_form_model(self, form_code):
        try:
            form_model = get_form_model_by_code(self.dbm, form_code)
        except FormModelDoesNotExistsException:
            form_model = get_form_model_by_code(self.dbm, "poll")
        return form_model

    def get_question_codes(self, form_code):
        form_model = self.select_form_model(form_code)
        question_codes = []
        form_fields = form_model.fields
        for aField in form_fields:
            question_codes.append(aField.code.lower())
        #if form_model.entity_type[0] == 'reporter' :
        #    question_codes.remove('eid')
        return question_codes, form_model


class KeyBasedSMSParser(SMSParser):
    MESSAGE_PREFIX = ur'^(\w+)\s+\.(\w+)\s+(\w+)'
    MESSAGE_TOKEN = ur"(\S+)(.*)"
    SEPARATOR = u" ."


    def _handle_tokens_with_only_separators(self, tokens):
        new_tokens = []
        for token in tokens:
            if is_empty(token): continue
            if is_empty("".join(token.split("."))): continue
            new_tokens.append(token.strip())
        return new_tokens

    def _parse_tokens(self, tokens, form_code):
        tokens = self._handle_tokens_with_only_separators(tokens)
        submission = OrderedDict()
        question_codes, form_model = self.get_question_codes(form_code)
        extra_data = []

        for token in tokens:
            if is_empty(token): continue
            field_code, answer = self._parse_token(token)
            if field_code in submission.keys():
                raise MultipleSubmissionsForSameCodeException(field_code)
            submission[field_code] = answer
            if field_code not in question_codes:
                extra_data.append(answer)
        return submission, extra_data

    def _parse_token(self, token):
        m = re.match(self.MESSAGE_TOKEN, token, flags=re.UNICODE)  # Match first non white space set of values.
        field_code, value = m.groups()
        return field_code.lower(), value.strip()


    def parse(self, message):
        assert is_string(message)
        try:
            form_code, tokens = self.form_code(message)
            submission, extra_data = self._parse_tokens(tokens, form_code)
        except SMSParserInvalidFormatException as ex:
            raise SMSParserInvalidFormatException(ex.data)
        except MultipleSubmissionsForSameCodeException as ex:
            raise MultipleSubmissionsForSameCodeException(ex.data[0])
        return form_code, submission, extra_data

    def form_code(self, message):
        message = self.clean(message)
        self.validate_format(self.MESSAGE_PREFIX, message)
        tokens = message.split(self.SEPARATOR)
        form_code = self.pop_form_code(tokens)
        return form_code, tokens


class OrderSMSParser(SMSParser):
    MESSAGE_PREFIX_FOR_ORDERED_SMS = ur'[^ ]+\s+[^ ]+'

    def __init__(self, dbm):
        self.dbm = dbm

    def _parse_ordered_tokens(self, tokens, question_codes, form_code):
        submission = OrderedDict()
        extra_data = []

        for token_index in range(len(tokens)):
            token = tokens[token_index]
            if is_empty(token): continue
            if len(question_codes) <= token_index:
                extra_data.append(token)
            else:
                submission[question_codes[token_index]] = token
        return submission, extra_data

    def parse(self, message):
        assert is_string(message)
        try:
            form_code, tokens = self.form_code(message)
            question_codes, form_model = self.get_question_codes(form_code)
            submission, extra_data = self._parse_ordered_tokens(tokens, question_codes, form_code)
        except SMSParserInvalidFormatException as ex:
            raise SMSParserInvalidFormatException(ex.data)
        return form_code, submission, extra_data

    def form_code(self, message):
        message = self.clean(message)
        self.validate_format(self.MESSAGE_PREFIX_FOR_ORDERED_SMS, message)
        tokens = message.split()
        return self.get_form_code_and_tokens(tokens)


class WebParser(object):
    def _remove_csrf_token(self, message):
        if 'csrfmiddlewaretoken' in message:
            message.pop('csrfmiddlewaretoken')

    def _fetch_string_value(self, message):
        return {code: self._to_str(value) for code, value in message.iteritems()}

    def parse(self, message):
        form_code = message.pop('form_code')
        self._remove_csrf_token(message)
        return form_code, self._fetch_string_value(message)

    def _to_str(self, value):
        if isinstance(value, (int, float, long)):
            return str(value)
        return "".join(value).strip() if value is not None else None


class CsvParser(object):
    EXTRA_VALUES = "extra_values"

    def _next_line(self, dict_reader):
        return dict_reader.respond().values()[0]

    def _parse_header(self, dict_reader):
        field_header = dict_reader.fieldnames

        if is_empty(field_header):
            raise CSVParserInvalidHeaderFormatException()

        self._remove_trailing_empty_header_field(field_header)

        if self._has_empty_values(field_header):
            raise CSVParserInvalidHeaderFormatException()

        return [field.strip().lower() for field in field_header]

    def _strip_field_values(self, row):
        for key, value in row.items():
            if value is not None and is_string(value):
                row[unicode(key, encoding='utf-8')] = unicode(value.strip(), encoding='utf-8')

    def _parse_row(self, form_code_fieldname, row):
        result_row = dict(row)
        self._strip_field_values(result_row)
        self._remove_extra_field_values(result_row)
        form_code = result_row.pop(form_code_fieldname).lower()
        return form_code, result_row

    def parse(self, csv_data):
        assert is_string(csv_data)
        csv_data = self._clean(csv_data)
        dict_reader = csv.DictReader(self._to_list(csv_data), restkey=self.EXTRA_VALUES)
        dict_reader.fieldnames = self._parse_header(dict_reader)
        parsedData = []
        form_code_fieldname = dict_reader.fieldnames[0]
        for row in dict_reader:
            parsedData.append(self._parse_row(form_code_fieldname, row))
        return parsedData

    def _has_empty_values(self, values_list):
        for value in values_list:
            if is_empty(value):
                return True
        return False

    def _remove_trailing_empty_header_field(self, field_header):
        for field in field_header[::-1]:
            if is_empty(field):
                field_header.pop()
            else:
                break

    def _remove_extra_field_values(self, result_row):
        if result_row.get(self.EXTRA_VALUES):
            result_row.pop(self.EXTRA_VALUES)

    def _clean(self, csv_data):
        return csv_data.strip()

    def _to_list(self, csv_data):
        return csv_data.splitlines()




class XlsParser(object):
    def parse(self, xls_contents):
        assert xls_contents is not None
        workbook = xlrd.open_workbook(file_contents=xls_contents)
        all_sheets = workbook.sheets()
        worksheet = self._get_worksheet(all_sheets)
        header_found = False
        header = None
        parsed_data = []
        for row_num in range(worksheet.nrows):
            row = worksheet.row_values(row_num)

            if not header_found:
                header, header_found = self._is_header_row(row)
                continue
            if self._is_empty(row):
                continue

            row = self._clean(row)
            row_dict = dict(zip(header, row))
            form_code, values = (row_dict.pop(header[0]).lower(), row_dict)
            parsed_data.append((form_code, values))
        if not header_found:
            raise XlsParserInvalidHeaderFormatException()
        return parsed_data

    def _get_worksheet(self, all_sheets):
        work_sheet = all_sheets[0]
        return all_sheets[1] if work_sheet.name == 'codes' else work_sheet

    def _get_code_sheet(self, all_sheets):
        work_sheet = all_sheets[0]
        return work_sheet if work_sheet.name == 'codes' else all_sheets[1]

    def _remove_trailing_empty_header_field(self, field_header):
        for field in field_header[::-1]:
            if is_empty(field):
                field_header.pop()
            else:
                break


    def _is_header_row(self, row):
        if is_empty(row[0]):
            return None, False
        self._remove_trailing_empty_header_field(row)
        return [unicode(value).strip().lower() for value in row], True

    def _clean(self, row):
        return [unicode(value).strip() for value in row]

    def _is_empty(self, row):
        return len([value for value in row if not is_empty(value)]) == 0

class XlsxParser(XlsParser):

    def parse(self, file_contents):
        assert file_contents is not None
        xlsx_file = StringIO.StringIO(file_contents)

        workbook = load_workbook(xlsx_file, use_iterators = True)
        worksheet = self._get_worksheet(workbook.worksheets)
        parsedData = []
        for row in worksheet.iter_rows():
            row_values = [self._get_value(x.value) for x in row]
            row_values = self._clean(row_values)
            parsedData.append(row_values)
        return parsedData
    
    def _get_value(self, value):
        if value is not None:
            return unicode(value)
        return ''
    def _get_worksheet(self, all_sheets):
        work_sheet = all_sheets[0]
        return all_sheets[1] if work_sheet.title == 'codes' else work_sheet

    def _get_code_sheet(self, all_sheets):
        work_sheet = all_sheets[0]
        return work_sheet if work_sheet.title == 'codes' else all_sheets[1]

class XlsOrderedParser(XlsParser):
    def parse(self, xls_contents):
        assert xls_contents is not None
        workbook = xlrd.open_workbook(file_contents=xls_contents)
        all_sheets = workbook.sheets()
        worksheet = self._get_worksheet(all_sheets)
        codes_sheet = self._get_code_sheet(all_sheets)
        row = codes_sheet.row_values(0)
        header, header_found = self._is_header_row(row)
        form_code = header[0]
        header = header[1:]
        parsed_data = []
        for row_num in range(1, worksheet.nrows):
            row = worksheet.row_values(row_num)
            row = self._clean(row)
            values = OrderedDict(zip(header, row))
            parsed_data.append((form_code, values))
        if not header_found:
            raise XlsParserInvalidHeaderFormatException()
        return parsed_data


class XFormParser(object):
    def __init__(self, dbm):
        self.dbm = dbm

    def is_field_set_answer(self, value):
        return type(value) is list

    def _fetch_string_value(self, message):
        str_dict = OrderedDict()
        for code, value in message.iteritems():
            if not self.is_field_set_answer(value):
                str_dict.update({code: self._to_str(value)})
            else:
                str_dict.update({code: value})

        return str_dict

    def parse(self, message):
        submission_dict = xmltodict.parse(message, 'utf-8').values()[0]
        # xform elements don't have namespace information since it's being default, so form_code don't include namespace
        form_code = submission_dict.pop('form_code')
        form_model = get_form_model_by_code(self.dbm, form_code)
        self.__format_response_fields(form_model, submission_dict)
        return form_code, self._fetch_string_value(submission_dict)

    def _to_str(self, value):
        if isinstance(value, (int, float, long)):
            return str(value)

        if type(value) is list:
            return " ".join(value)

        return value if value is not None else None

    def _format_field_set(self, field, values):
        if type(values) is not list:
            values = [values]

        for field in field.fields:
            [self._format_field(field, value) for value in values]

        return values

    def _parse_time_string(self, time_string):
        return time_string.split(":00.000")[0] if time_string else time_string

    def _format_field(self, field, values):
        code = field.code
        if not values.get(code):
            return
        if type(field) == GeoCodeField:
            geo_code_list = values[code].split(' ')
            values[code] = '{0},{1}'.format(geo_code_list[0], geo_code_list[1])
        if type(field) == DateField:
            date = datetime.strptime(values[code], "%Y-%m-%d")
            values[code] = self.__get_formatted_date(field.date_format, date)
        if type(field) == IntegerField:
            values[code] = "%s" % values[code].__str__()
        if type(field) == TimeField:
            values[code] = "%s" % self._parse_time_string(values[code])
        if type(field) == DateTimeField:
            values[code] = self._parse_date_time(values[code])
        if type(field) == FieldSet:
            value_list = self._format_field_set(field, values[code])
            values[code] =  [self._fetch_string_value(r) for r in value_list]
        if isinstance(field, MediaField):
            values[code] = self._get_file_name(values[code])

    def _get_file_name(self, value):
        web_new_file_selected = type(value) is OrderedDict
        return [v for k,v in value.iteritems() if k == '#text'][0] if web_new_file_selected\
                else value

    def __format_response_fields(self, form_model, values):
        [self._format_field(field, values) for field in form_model.fields]
        return values

    def __get_formatted_date(self, date_format, date):
        return format_date(date, DateField.FORMAT_DATE_DICTIONARY.get(date_format))

    def _get_attachment(self, field):
        return field['#text']

    def _parse_date_time(self, value):
        # 2015-01-13T21:45:00.000+06:30
        # Remove timezone and milli seconds.
        date_time_without_milliseconds = re.compile("\.\d{3}\+" ).split(value)[0]
        return datetime.strptime(date_time_without_milliseconds, '%Y-%m-%dT%H:%M:%S').strftime('%d.%m.%Y %H:%M:%S')

class XlsDatasenderParser(XlsParser):
    def parse(self, xls_contents):
        assert xls_contents is not None
        workbook = xlrd.open_workbook(file_contents=xls_contents)
        worksheet = self._get_worksheet(workbook.sheets())
        codes_sheet = self._get_code_sheet(workbook.sheets())
        parsed_data = []
        row = codes_sheet.row_values(0)
        header, header_found = self._is_header_row(row)

        if row[0] != 'reg':
            raise Exception("Invalid datasender excel imported")

        form_code = REGISTRATION_FORM_CODE
        header = header[1:]
        for row_num in range(1, worksheet.nrows):
            row = worksheet.row_values(row_num)
            row = self._clean(row)
            values = dict(zip(header, row))
            values.update({"t": "reporter"})
            parsed_data.append((form_code, values))
        if not header_found:
            raise XlsParserInvalidHeaderFormatException()
        return parsed_data

class XlsxDataSenderParser(XlsxParser):

    def parse(self, file_contents):
        assert file_contents is not None
        xlsx_file = StringIO.StringIO(file_contents)

        workbook = load_workbook(xlsx_file, use_iterators = True)
        codes_sheet = workbook.get_sheet_by_name('codes')
        worksheet = self._get_worksheet(workbook.worksheets)

        rows = []
        for cs in codes_sheet.iter_rows():
            rows = [self._get_value(x.value) for x in cs]
        header, header_found = self._is_header_row(rows)
        parsed_data = []
        form_code = REGISTRATION_FORM_CODE
        header = header[1:]
        for row in worksheet.iter_rows(row_offset=1):
            row_values = [self._get_value(x.value) for x in row]
            values = dict(zip(header, row_values))
            values.update({"t": "reporter"})
            parsed_data.append((form_code, values))
        if not header_found:
            raise XlsParserInvalidHeaderFormatException()
        return parsed_data

